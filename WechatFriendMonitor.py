#! python3
#coding:utf-8

######################################
# @program_name: spider.py
#	@author：Jessus
#	@create_date: 2017-08-23
#	@modifyer: Jessus
#	@modify_date: 2017-08-23
#	@features: 监控微信公众号文章的阅读量和点赞量
######################################



import sys, time, re

#Excel操作模块
import openpyxl

#网络访问类模块
import urllib,json,requests
from urllib.request import quote

#邮件类模块
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header

#微信机器人模块
import itchat
from itchat.content import *

#多线程模块
import threading


class WechatArticleSpider:
	#1、构造函数
	def __init__(self,url,interval,repeat_times):
		self.url=url
		self.interval=interval
		self.repeat_times=repeat_times
			
		info_appid='483d5dc444434814c5b58c67c575a491'
		
		self.article_dict={}
		article_url=quote(url)
		info_api_url='http://api.shenjianshou.cn/?appid=' + info_appid + '&url=' + article_url
		resp_info=requests.get(info_api_url)
		
		if eval(resp_info.text)['error_code'] == 0: #判断接口执行返回码，0为正确，其他为失败
			article_title=eval(resp_info.text)['data']['article_title']
			article_author=eval(resp_info.text)['data']['article_author']
			article_publish_time=time.strftime('%Y%m%d',time.localtime(int(eval(resp_info.text)['data']['article_publish_time'])))
			article_fixed_url=eval(resp_info.text)['data']['article_fixed_url']
			
			self.article_dict['article_title']=article_title
			self.article_dict['article_author']=article_author
			self.article_dict['article_publish_time']=article_publish_time
			self.article_dict['article_fixed_url']=article_fixed_url
			self.article_dict['article_view_count']={}
			self.article_dict['article_agree_count']={}
		else: #打印接口失败原因
			print('***' + time.strftime('%Y%m%d %H:%M:%S',time.localtime(time.time())) + '文章信息接口错误:' + eval(resp_comment.text)['reason'])
			print('***' + time.strftime('%Y%m%d %H:%M:%S',time.localtime(time.time())) + '请输入<Ctrl+c>终止机器人继续执行!!!!!!!!!!!!!!!!!')
			
		
	def articlegetcomment(self,article_fixed_url):
		#获取文章阅读和点赞量收费接口URL构造
		comment_appid='330908509ce59a903df7f3ac53669d28'
		article_fixed_url=quote(article_fixed_url)
		comment_api_url='http://api.shenjianshou.cn/?appid=' + comment_appid + '&url=' + article_fixed_url
		
		for i in range(self.repeat_times): #按重复次数执行监控
			op_time=time.strftime('%Y%m%d%H%M%S',time.localtime(time.time()))
			resp_comment=requests.get(comment_api_url) #调用文章阅读和点赞量接口       
			if eval(resp_comment.text)['error_code'] == 0: #判断接口执行返回码，0为正确，其他为失败
				article_view_count=eval(resp_comment.text)['data']['article_view_count']
				article_agree_count=eval(resp_comment.text)['data']['article_agree_count']
				self.article_dict['article_view_count'][op_time]=article_view_count
				self.article_dict['article_agree_count'][op_time]=article_agree_count
				time.sleep(self.interval*60) #按间隔时间进行等待
			else:  #打印接口失败原因
				print('***' + time.strftime('%Y%m%d %H:%M:%S',time.localtime(time.time())) + '文章阅读量和点赞量接口错误:' + eval(resp_comment.text)['reason'])
				print('***' + time.strftime('%Y%m%d %H:%M:%S',time.localtime(time.time())) + '请输入<Ctrl+c>终止机器人继续执行!!!!!!!!!!!!!!!!!')
		return
		
	def outputexcel(self):
		wb=openpyxl.Workbook()
		sheet=wb.active
		sheet.title='Article'
		row_num=1
		column_num=1
		for item in self.article_dict.keys():
			if item == 'article_view_count':
				sheet.cell(row=row_num,column=column_num).value = '文章阅读量'
				sheet.cell(row=row_num,column=column_num+1).value	= 'time'
				sheet.cell(row=row_num,column=column_num+2).value	= 'view_count'
				row_num+=1
					
				for i in self.article_dict[item].keys():
					sheet.cell(row=row_num,column=column_num+1).value	= i
					sheet.cell(row=row_num,column=column_num+2).value	= int(self.article_dict[item][i])
					row_num+=1
					
				#Excel自动画图，View_Count	
				data = openpyxl.chart.Reference(sheet,min_col=column_num+2,min_row=row_num-len(self.article_dict[item])-1,max_col=column_num+2,max_row=row_num-1)
				label = openpyxl.chart.Reference(sheet,min_col=column_num+1,min_row=row_num-len(self.article_dict[item]),max_col=column_num+1,max_row=row_num)
				chartObj=openpyxl.chart.LineChart()
				chartObj.add_data(data,titles_from_data=True)
				chartObj.set_categories(label)
				sheet.add_chart(chartObj,'B'+str(row_num + len(self.article_dict[item].keys()) + 2))		#View_count 图定位					
			elif item == 'article_agree_count':
				sheet.cell(row=row_num,column=column_num).value = '文章点赞量'
				sheet.cell(row=row_num,column=column_num+1).value	= 'time'
				sheet.cell(row=row_num,column=column_num+2).value	= 'agree_count'
				row_num+=1
					
				for i in self.article_dict[item].keys():
					sheet.cell(row=row_num,column=column_num+1).value	= i
					sheet.cell(row=row_num,column=column_num+2).value	= int(self.article_dict[item][i])
					row_num+=1
				
				#Excel画图，agree_count
				data2 = openpyxl.chart.Reference(sheet,min_col=column_num+2,min_row=row_num-len(self.article_dict[item])-1,max_col=column_num+2,max_row=row_num-1)
				label2 = openpyxl.chart.Reference(sheet,min_col=column_num+1,min_row=row_num-len(self.article_dict[item]),max_col=column_num+1,max_row=row_num)
				chartObj2=openpyxl.chart.LineChart()
				chartObj2.add_data(data2,titles_from_data=True)
				chartObj2.set_categories(label2)
				sheet.add_chart(chartObj2,'L'+str(row_num+1))	#Agree_count 图定位				
			else:
				sheet.cell(row=row_num,column=column_num).value = item
				sheet.cell(row=row_num,column=column_num+1).value = str(self.article_dict[item])
				row_num+=1

		wb.save('Monitor_Wechat_Article.xlsx')
	
	#将结果发邮件的方法
	def sendemail(self):
		sender = '13688880130@139.com'  #定义邮件发送者
		receivers = ['13688880130@139.com',\
		             'jianglufeng@gzliuhe.com']  # 接收邮件，可设置为你的QQ邮箱或者其他邮箱
		 
		#创建一个带附件的实例
		message = MIMEMultipart()
		message['From'] = Header("六合“中国移动10086”公众号监控机器人", 'utf-8')
		message['To'] =  Header("中移在线广东媒体监控组", 'utf-8')
		subject = '“中国移动10086”公众号文章监控'
		message['Subject'] = Header(subject, 'utf-8')
		 
		#邮件正文内容
		message.attach(MIMEText('微信公众号文章监控结果……', 'plain', 'utf-8'))
		 
		# 构造附件1，传送当前目录下的 Monitor_Wechat_Article.xlsx 文件
		att1 = MIMEText(open('Monitor_Wechat_Article.xlsx', 'rb').read(), 'base64', 'utf-8')
		att1["Content-Type"] = 'application/octet-stream'
		# 这里的filename可以任意写，写什么名字，邮件中显示什么名字
		att1["Content-Disposition"] = 'attachment; filename="Monitor_Wechat_Article.xlsx"'
		message.attach(att1)
		 
		#### 构造附件2
		###att2 = MIMEText(open('spider.py', 'rb').read(), 'base64', 'utf-8')
		###att2["Content-Type"] = 'application/octet-stream'
		###att2["Content-Disposition"] = 'attachment; filename="spider.py"'
		###message.attach(att2)
		 
		try:
			smtp_server=smtplib.SMTP_SSL('smtp.139.com',465)
			smtp_server.login('13688880130@139.com','newlife@2012')
			smtp_server.sendmail(sender, receivers, message.as_string())
		except smtplib.SMTPException:
		    print ("Error: 无法发送邮件")
	
	
	def run(self):	
		###多线程同时监控多篇文章
		##threads=[]
		##for articleIdx in self.article_dict.keys():
		##	t=threading.Thread(target=self.articlegetcomment,args=(articleIdx,self.article_dict[articleIdx]['article_fixed_url'],))
		##	threads.append(t)
		##	
		##for t in threads:
		##	t.setDaemon(True)
		##	t.start()
		##
		##t.join()
		
		#由于改用itchat模拟监控公众号，因此每次只能获得最新发布的一篇文章消息，因此不再需要多线程，但保留多线程实现代码			
		self.articlegetcomment(self.article_dict['article_fixed_url'])
		self.outputexcel()
		self.sendemail()
		print('***' + time.strftime('%Y%m%d %H:%M:%S',time.localtime(time.time())) + ':  ' + self.article_dict['article_title'] +' 监控结束，监控结果已发邮件..............')
		return

#---------------End of Class WechatArticleSpider

#微信机器人动作函数
@itchat.msg_register(SHARING,isFriendChat=True)
def catch_sharing(msg):	
	url=msg['Url']
	urlCheckRegex=re.compile(r'(http://mp.weixin.qq.com/)(\w*)')
	urlCheckMo=urlCheckRegex.search(url)
	if urlCheckMo != None:  
		print('***' + time.strftime('%Y%m%d %H:%M:%S',time.localtime(time.time())) + ':  已捕获“中国移动10086”公众号发出文章，文章名：' + msg['FileName'])
		print('***' + time.strftime('%Y%m%d %H:%M:%S',time.localtime(time.time())) + ':  文章：' + msg['FileName'] +' 启动监控..............')
		WechatArticleSpider(url,1,3).run()


#微信登录成功后执行的动作
def lc():
	print('##################################################')
	print('#\t中国移动10086公众号监控机器人')
	print('#\t提供者：广州六合信息科技股份有限公司')
	print('##################################################')
	print('***' + time.strftime('%Y%m%d %H:%M:%S',time.localtime(time.time())) + ':  开始监控“中国移动10086”公众号...........按<Ctrl+c)退出程序')

#微信退出登录后执行的动作	
def ec():
	print('##################################################')
	print('\t\t#中国移动10086公众号监控')
	print('\t\t#提供者：广州六合信息科技股份有限公司')
	print('##################################################')

#执行部分
if __name__ == '__main__':
	itchat.auto_login(loginCallback=lc,exitCallback=ec,hotReload=True)
	
	try:
		itchat.run()
	except KeyboardInterrupt:
		itchat.logout()
	
	
	



